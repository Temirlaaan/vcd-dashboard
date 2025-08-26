# export_service.py - Сервис для экспорта данных
import io
import os
from datetime import datetime
from typing import List, Dict, Optional
import pandas as pd
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils.dataframe import dataframe_to_rows
from openpyxl.chart import BarChart, PieChart, Reference
from reportlab.lib import colors
from reportlab.lib.pagesizes import A4, landscape
from reportlab.lib.styles import getSampleStyleSheet, ParagraphStyle
from reportlab.lib.units import inch
from reportlab.platypus import SimpleDocTemplate, Table, TableStyle, Paragraph, Spacer, PageBreak, Image
from reportlab.graphics.shapes import Drawing
from reportlab.graphics.charts.piecharts import Pie
from reportlab.graphics.charts.barcharts import VerticalBarChart
from reportlab.lib.enums import TA_CENTER, TA_LEFT

class ExportService:
    """Сервис для экспорта данных в различные форматы"""
    
    def __init__(self):
        self.styles = getSampleStyleSheet()
        self._setup_pdf_styles()
    
    def _setup_pdf_styles(self):
        """Настройка стилей для PDF"""
        self.styles.add(ParagraphStyle(
            name='CustomTitle',
            parent=self.styles['Heading1'],
            fontSize=24,
            textColor=colors.HexColor('#1e293b'),
            alignment=TA_CENTER,
            spaceAfter=30
        ))
        
        self.styles.add(ParagraphStyle(
            name='CustomHeading',
            parent=self.styles['Heading2'],
            fontSize=16,
            textColor=colors.HexColor('#334155'),
            spaceBefore=20,
            spaceAfter=10
        ))
        
        self.styles.add(ParagraphStyle(
            name='InfoText',
            parent=self.styles['Normal'],
            fontSize=10,
            textColor=colors.HexColor('#64748b')
        ))
    
    def export_to_excel(
        self,
        dashboard_data: Dict,
        history: Optional[List[Dict]] = None,
        notes: Optional[List[Dict]] = None,
        filename: Optional[str] = None
    ) -> bytes:
        """Экспорт данных в Excel"""
        
        # Создаем новый workbook
        wb = Workbook()
        
        # Удаляем default sheet
        wb.remove(wb.active)
        
        # Стили
        header_font = Font(bold=True, color="FFFFFF")
        header_fill = PatternFill(start_color="6366F1", end_color="6366F1", fill_type="solid")
        header_alignment = Alignment(horizontal="center", vertical="center")
        border = Border(
            left=Side(style='thin'),
            right=Side(style='thin'),
            top=Side(style='thin'),
            bottom=Side(style='thin')
        )
        
        # 1. Общая статистика
        ws_summary = wb.create_sheet("Summary")
        ws_summary.append(["VCD IP Manager Report"])
        ws_summary.append([f"Generated: {datetime.now().strftime('%Y-%m-%d %H:%M:%S')}"])
        ws_summary.append([])
        ws_summary.append(["Metric", "Value"])
        ws_summary.append(["Total Clouds", dashboard_data['total_clouds']])
        ws_summary.append(["Total IPs", dashboard_data['total_ips']])
        ws_summary.append(["Used IPs", dashboard_data['used_ips']])
        ws_summary.append(["Free IPs", dashboard_data['free_ips']])
        ws_summary.append(["Usage %", f"{dashboard_data['usage_percentage']}%"])
        
        # Форматирование заголовка
        ws_summary['A1'].font = Font(bold=True, size=16)
        ws_summary.merge_cells('A1:B1')
        ws_summary['A4'].font = header_font
        ws_summary['A4'].fill = header_fill
        ws_summary['B4'].font = header_font
        ws_summary['B4'].fill = header_fill
        
        # 2. Детали по облакам
        ws_clouds = wb.create_sheet("Cloud Details")
        cloud_headers = ["Cloud", "Pools", "Total IPs", "Used IPs", "Free IPs", "Usage %"]
        ws_clouds.append(cloud_headers)
        
        for cloud in dashboard_data['clouds']:
            ws_clouds.append([
                cloud['cloud_name'].upper(),
                cloud['total_pools'],
                cloud['total_ips'],
                cloud['used_ips'],
                cloud['free_ips'],
                f"{cloud['usage_percentage']}%"
            ])
        
        # Форматирование заголовков
        for cell in ws_clouds[1]:
            cell.font = header_font
            cell.fill = header_fill
            cell.alignment = header_alignment
        
        # 3. Детали по пулам
        ws_pools = wb.create_sheet("Pool Details")
        pool_headers = ["Cloud", "Pool Name", "Network", "Total", "Used", "Free", "Usage %"]
        ws_pools.append(pool_headers)
        
        for cloud in dashboard_data['clouds']:
            for pool in cloud['pools']:
                ws_pools.append([
                    cloud['cloud_name'].upper(),
                    pool['name'],
                    pool['network'],
                    pool['total_ips'],
                    pool['used_ips'],
                    pool['free_ips'],
                    f"{pool['usage_percentage']}%"
                ])
        
        # Форматирование заголовков
        for cell in ws_pools[1]:
            cell.font = header_font
            cell.fill = header_fill
            cell.alignment = header_alignment
        
        # 4. Занятые IP адреса
        ws_allocated = wb.create_sheet("Allocated IPs")
        allocated_headers = ["IP Address", "Organization", "Cloud", "Pool", "Type", "Entity"]
        ws_allocated.append(allocated_headers)
        
        for alloc in dashboard_data['all_allocations']:
            ws_allocated.append([
                alloc['ip_address'],
                alloc['org_name'],
                alloc['cloud_name'].upper(),
                alloc['pool_name'],
                alloc['allocation_type'],
                alloc.get('entity_name', '-')
            ])
        
        # Форматирование заголовков
        for cell in ws_allocated[1]:
            cell.font = header_font
            cell.fill = header_fill
            cell.alignment = header_alignment
        
        # 5. История (если есть)
        if history:
            ws_history = wb.create_sheet("History")
            history_headers = ["Timestamp", "Action", "User", "IP", "Cloud", "Details"]
            ws_history.append(history_headers)
            
            for entry in history:
                ws_history.append([
                    entry['timestamp'],
                    entry['action_type'],
                    entry['user'],
                    entry.get('ip_address', '-'),
                    entry.get('cloud_name', '-'),
                    entry.get('details', '-')
                ])
            
            # Форматирование заголовков
            for cell in ws_history[1]:
                cell.font = header_font
                cell.fill = header_fill
                cell.alignment = header_alignment
        
        # 6. Заметки (если есть)
        if notes:
            ws_notes = wb.create_sheet("Notes")
            notes_headers = ["Created", "Author", "Title", "Category", "Priority", "Content"]
            ws_notes.append(notes_headers)
            
            for note in notes:
                ws_notes.append([
                    note['created_at'],
                    note['author'],
                    note['title'],
                    note['category'],
                    note['priority'],
                    note['content'][:100] + '...' if len(note['content']) > 100 else note['content']
                ])
            
            # Форматирование заголовков
            for cell in ws_notes[1]:
                cell.font = header_font
                cell.fill = header_fill
                cell.alignment = header_alignment
        
        # Автоподбор ширины колонок для всех листов
        for ws in wb.worksheets:
            for column in ws.columns:
                max_length = 0
                column_letter = column[0].column_letter
                for cell in column:
                    try:
                        if len(str(cell.value)) > max_length:
                            max_length = len(str(cell.value))
                    except:
                        pass
                adjusted_width = min(max_length + 2, 50)
                ws.column_dimensions[column_letter].width = adjusted_width
        
        # Сохраняем в bytes
        excel_file = io.BytesIO()
        wb.save(excel_file)
        excel_file.seek(0)
        
        return excel_file.getvalue()
    
    def export_to_pdf(
        self,
        dashboard_data: Dict,
        history: Optional[List[Dict]] = None,
        notes: Optional[List[Dict]] = None,
        filename: Optional[str] = None
    ) -> bytes:
        """Экспорт данных в PDF"""
        
        # Создаем буфер для PDF
        pdf_buffer = io.BytesIO()
        
        # Создаем документ
        doc = SimpleDocTemplate(
            pdf_buffer,
            pagesize=landscape(A4),
            rightMargin=30,
            leftMargin=30,
            topMargin=30,
            bottomMargin=30
        )
        
        # Элементы для документа
        elements = []
        
        # Заголовок
        title = Paragraph("VCD IP Manager Report", self.styles['CustomTitle'])
        elements.append(title)
        
        timestamp = Paragraph(
            f"Generated: {datetime.now().strftime('%Y-%m-%d %H:%M:%S')}",
            self.styles['InfoText']
        )
        elements.append(timestamp)
        elements.append(Spacer(1, 20))
        
        # 1. Общая статистика
        elements.append(Paragraph("Executive Summary", self.styles['CustomHeading']))
        
        summary_data = [
            ['Metric', 'Value'],
            ['Total Clouds', str(dashboard_data['total_clouds'])],
            ['Total IP Addresses', str(dashboard_data['total_ips'])],
            ['Used IPs', str(dashboard_data['used_ips'])],
            ['Free IPs', str(dashboard_data['free_ips'])],
            ['Overall Usage', f"{dashboard_data['usage_percentage']}%"]
        ]
        
        summary_table = Table(summary_data, colWidths=[3*inch, 2*inch])
        summary_table.setStyle(TableStyle([
            ('BACKGROUND', (0, 0), (-1, 0), colors.HexColor('#6366F1')),
            ('TEXTCOLOR', (0, 0), (-1, 0), colors.whitesmoke),
            ('ALIGN', (0, 0), (-1, -1), 'CENTER'),
            ('FONTNAME', (0, 0), (-1, 0), 'Helvetica-Bold'),
            ('FONTSIZE', (0, 0), (-1, 0), 12),
            ('BOTTOMPADDING', (0, 0), (-1, 0), 12),
            ('BACKGROUND', (0, 1), (-1, -1), colors.beige),
            ('GRID', (0, 0), (-1, -1), 1, colors.black)
        ]))
        
        elements.append(summary_table)
        elements.append(Spacer(1, 20))
        
        # 2. График использования по облакам (простая таблица вместо графика)
        elements.append(Paragraph("Cloud Usage Overview", self.styles['CustomHeading']))
        
        cloud_data = [['Cloud', 'Pools', 'Total IPs', 'Used', 'Free', 'Usage %']]
        for cloud in dashboard_data['clouds']:
            cloud_data.append([
                cloud['cloud_name'].upper(),
                str(cloud['total_pools']),
                str(cloud['total_ips']),
                str(cloud['used_ips']),
                str(cloud['free_ips']),
                f"{cloud['usage_percentage']}%"
            ])
        
        cloud_table = Table(cloud_data, colWidths=[1.5*inch, 1*inch, 1.5*inch, 1*inch, 1*inch, 1*inch])
        cloud_table.setStyle(TableStyle([
            ('BACKGROUND', (0, 0), (-1, 0), colors.HexColor('#6366F1')),
            ('TEXTCOLOR', (0, 0), (-1, 0), colors.whitesmoke),
            ('ALIGN', (0, 0), (-1, -1), 'CENTER'),
            ('FONTNAME', (0, 0), (-1, 0), 'Helvetica-Bold'),
            ('FONTSIZE', (0, 0), (-1, 0), 11),
            ('BOTTOMPADDING', (0, 0), (-1, 0), 12),
            ('BACKGROUND', (0, 1), (-1, -1), colors.lightgrey),
            ('GRID', (0, 0), (-1, -1), 1, colors.black)
        ]))
        
        elements.append(cloud_table)
        elements.append(PageBreak())
        
        # 3. Детали по пулам
        elements.append(Paragraph("IP Pool Details", self.styles['CustomHeading']))
        
        # Разбиваем пулы на страницы по 10 штук
        for cloud in dashboard_data['clouds']:
            elements.append(Paragraph(f"{cloud['cloud_name'].upper()} Pools:", self.styles['Heading3']))
            
            pool_data = [['Pool Name', 'Network', 'Total', 'Used', 'Free', 'Usage']]
            for pool in cloud['pools']:
                usage = pool['usage_percentage']
                # Цветовая индикация использования
                if usage > 80:
                    usage_str = f"⚠️ {usage}%"
                elif usage > 50:
                    usage_str = f"⚡ {usage}%"
                else:
                    usage_str = f"✅ {usage}%"
                
                pool_data.append([
                    pool['name'][:30],  # Обрезаем длинные имена
                    pool['network'],
                    str(pool['total_ips']),
                    str(pool['used_ips']),
                    str(pool['free_ips']),
                    usage_str
                ])
            
            pool_table = Table(pool_data, colWidths=[2.5*inch, 1.8*inch, 0.8*inch, 0.8*inch, 0.8*inch, 1*inch])
            pool_table.setStyle(TableStyle([
                ('BACKGROUND', (0, 0), (-1, 0), colors.HexColor('#8B5CF6')),
                ('TEXTCOLOR', (0, 0), (-1, 0), colors.whitesmoke),
                ('ALIGN', (0, 0), (-1, -1), 'CENTER'),
                ('FONTNAME', (0, 0), (-1, 0), 'Helvetica-Bold'),
                ('FONTSIZE', (0, 0), (-1, 0), 10),
                ('BOTTOMPADDING', (0, 0), (-1, 0), 8),
                ('BACKGROUND', (0, 1), (-1, -1), colors.lightblue),
                ('GRID', (0, 0), (-1, -1), 1, colors.black),
                ('FONTSIZE', (0, 1), (-1, -1), 9),
            ]))
            
            elements.append(pool_table)
            elements.append(Spacer(1, 15))
        
        # 4. Последние записи истории (если есть)
        if history and len(history) > 0:
            elements.append(PageBreak())
            elements.append(Paragraph("Recent Activity History", self.styles['CustomHeading']))
            
            # Показываем только последние 20 записей
            recent_history = history[:20]
            history_data = [['Time', 'Action', 'User', 'IP', 'Details']]
            
            for entry in recent_history:
                history_data.append([
                    entry['timestamp'][:19],  # Обрезаем микросекунды
                    entry['action_type'].replace('_', ' '),
                    entry['user'],
                    entry.get('ip_address', '-'),
                    (entry.get('details', '-')[:40] + '...') if entry.get('details') and len(entry.get('details', '')) > 40 else entry.get('details', '-')
                ])
            
            history_table = Table(history_data, colWidths=[2*inch, 1.5*inch, 1.5*inch, 1.5*inch, 2.5*inch])
            history_table.setStyle(TableStyle([
                ('BACKGROUND', (0, 0), (-1, 0), colors.HexColor('#10B981')),
                ('TEXTCOLOR', (0, 0), (-1, 0), colors.whitesmoke),
                ('ALIGN', (0, 0), (-1, -1), 'LEFT'),
                ('FONTNAME', (0, 0), (-1, 0), 'Helvetica-Bold'),
                ('FONTSIZE', (0, 0), (-1, 0), 10),
                ('BOTTOMPADDING', (0, 0), (-1, 0), 8),
                ('BACKGROUND', (0, 1), (-1, -1), colors.lightgreen),
                ('GRID', (0, 0), (-1, -1), 1, colors.black),
                ('FONTSIZE', (0, 1), (-1, -1), 8),
            ]))
            
            elements.append(history_table)
        
        # 5. Важные заметки (если есть)
        if notes and len(notes) > 0:
            elements.append(PageBreak())
            elements.append(Paragraph("Important Notes", self.styles['CustomHeading']))
            
            # Показываем только закрепленные или важные заметки
            important_notes = [n for n in notes if n.get('is_pinned') or n.get('priority') in ['high', 'critical']][:10]
            
            if important_notes:
                for note in important_notes:
                    note_title = f"📌 {note['title']}" if note.get('is_pinned') else note['title']
                    
                    if note.get('priority') == 'critical':
                        note_title = f"🔴 {note_title}"
                    elif note.get('priority') == 'high':
                        note_title = f"🟠 {note_title}"
                    
                    elements.append(Paragraph(note_title, self.styles['Heading4']))
                    elements.append(Paragraph(f"By {note['author']} on {note['created_at'][:10]}", self.styles['InfoText']))
                    
                    # Обрезаем длинный контент
                    content = note['content'][:300] + '...' if len(note['content']) > 300 else note['content']
                    elements.append(Paragraph(content, self.styles['Normal']))
                    elements.append(Spacer(1, 10))
        
        # Генерируем PDF
        doc.build(elements)
        pdf_buffer.seek(0)
        
        return pdf_buffer.getvalue()

# Глобальный экземпляр сервиса
export_service = ExportService()